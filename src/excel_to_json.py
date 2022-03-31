import openpyxl
import json
import streamlit as st

def excel_to_json(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    json_path = filename.split(".", 1)[0] + ".json"
    row = sheet.max_row
    data = []
    for r in range(2, row + 1):
        attr = {
            'Name': sheet.cell(row=r, column=2).value,
            'Skill': sheet.cell(row=r, column=3).value,
            'Skill Level': sheet.cell(row=r, column=4).value
        }
        data.append(attr)
    
    with open(json_path, mode = 'w', encoding = 'utf-8') as f:
        f.write(json.dumps(data, ensure_ascii = False, indent = 4))

filename = "hr_sample_f.xlsx"
excel_to_json(filename)